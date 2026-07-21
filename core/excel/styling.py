"""
core/excel/styling.py — visual polish for DynamicSchema workbooks: header
styling, borders, number formats, dropdown data validation, and charts.
Kept separate from schema_dynamic.py/derived_sheets.py (which only ever
deal with plain values) so "how it looks" is one place, independent of
"what the numbers are" — see feedback_excel_presentable_mobile.md (must
be readable opened straight from a phone) and feedback_modular_structure.md.

Colors/formats match the real Finances_2026_v2.xlsx file (the user's own
named style reference), confirmed by direct structural inspection, not
guessed: dark-navy bold-white header, thin borders, dd.mm.yyyy dates,
#,##0.00 amounts.
"""
from __future__ import annotations

from openpyxl.chart import LineChart, PieChart, Reference
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .template_model import ROLE_AMOUNT, ROLE_BASE_AMOUNT, ROLE_CATEGORY, ROLE_CURRENCY, ROLE_DATE, ROLE_PAYMENT, ROLE_RATE, ROLE_TYPE

HEADER_FONT = Font(bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="FF1C1C30")
BORDER = Border(*(Side(style="thin", color="FFD0D0D0") for _ in range(4)))
DATE_FORMAT = "dd.mm.yyyy"
AMOUNT_FORMAT = "#,##0.00"
MAX_STYLED_ROW = 100

_ROLE_WIDTH = {
    ROLE_DATE: 12, ROLE_TYPE: 11, ROLE_CATEGORY: 15, ROLE_PAYMENT: 11, ROLE_CURRENCY: 10,
    ROLE_AMOUNT: 12, ROLE_RATE: 10, ROLE_BASE_AMOUNT: 15,
}
_LIST_SOURCE = {ROLE_TYPE: ("B", 20), ROLE_CATEGORY: ("A", 100), ROLE_PAYMENT: ("C", 20), ROLE_CURRENCY: ("D", 20)}


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill = HEADER_FONT, HEADER_FILL
    ws.freeze_panes = "A2"


def style_transaction_sheet(ws, col: dict[str, int], max_row: int = MAX_STYLED_ROW) -> None:
    """Header + borders + number formats + column widths for a month
    sheet or the All Transactions sheet -- both share the same role->column
    layout, so one function styles either."""
    ncols = max(col.values())
    _style_header(ws, ncols)
    date_col, rate_col, base_col = col.get(ROLE_DATE), col.get(ROLE_RATE), col.get(ROLE_BASE_AMOUNT)
    amount_cols = {c for c in (col.get(ROLE_AMOUNT), rate_col, base_col) if c}
    for row in range(2, max_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).border = BORDER
        if date_col:
            ws.cell(row=row, column=date_col).number_format = DATE_FORMAT
        for c in amount_cols:
            ws.cell(row=row, column=c).number_format = AMOUNT_FORMAT
    for role, width in _ROLE_WIDTH.items():
        if role in col:
            ws.column_dimensions[get_column_letter(col[role])].width = width


def style_grid_sheet(ws, ncols: int, max_row: int, amount_from_col: int = 2) -> None:
    """Header + borders + number formats for a summary-shaped sheet
    (Monthly/Annual Summary, By Category) -- column 1 is a label
    (month/year/category name), everything from `amount_from_col` on is a
    number."""
    _style_header(ws, ncols)
    ws.column_dimensions["A"].width = 16
    for c in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    for row in range(2, max_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER
            if c >= amount_from_col:
                cell.number_format = AMOUNT_FORMAT


def effective_amount_col(col: dict[str, int]) -> int | None:
    """Which column holds the (hidden) effective-base-amount helper for
    this role map, or None if the template has no currency tracking at
    all -- the single source of truth for the helper's position, shared
    by add_effective_amount_column() (writes it) and derived_sheets.py
    (references it in SUMIFS formulas) so the two can never drift apart."""
    return max(col.values()) + 1 if ROLE_CURRENCY in col else None


def add_effective_amount_column(ws, col: dict[str, int], max_row: int = MAX_STYLED_ROW) -> None:
    """A hidden per-row formula column mirroring what
    DynamicSchema.convert_transaction() would resolve to in Python --
    prefers a persisted Amount(base currency) cell when the row has one,
    else falls back to the *current* Lists!H:I rate table (the same
    fallback resolve_rate() itself uses when no historical rate_history.json
    entry exists -- Excel formulas have no access to that local cache, so
    this is deliberately the "current table" half of that fallback only).
    Lets Monthly Summary/By Category sum a phone-entered row -- one with
    no Rate/Amount(base) filled in at all -- just as correctly as an
    app-written one. No-op if this template has no Currency column."""
    helper_col = effective_amount_col(col)
    if helper_col is None:
        return
    amount_letter = get_column_letter(col[ROLE_AMOUNT])
    currency_letter = get_column_letter(col[ROLE_CURRENCY])
    base_letter = get_column_letter(col[ROLE_BASE_AMOUNT]) if ROLE_BASE_AMOUNT in col else None
    for row in range(2, max_row + 1):
        rate_lookup = f"{amount_letter}{row}*IFERROR(VLOOKUP({currency_letter}{row},Lists!$H:$I,2,0),1)"
        if base_letter:
            formula = f'=IF({base_letter}{row}<>"",{base_letter}{row},{rate_lookup})'
        else:
            formula = f"={rate_lookup}"
        ws.cell(row=row, column=helper_col).value = formula
    letter = get_column_letter(helper_col)
    ws.column_dimensions[letter].hidden = True


def add_dropdowns(ws, col: dict[str, int], max_row: int = MAX_STYLED_ROW, lists_sheet: str = "Lists") -> None:
    """Data-validation dropdown pick-lists for Type/Category/Payment/
    Currency, sourced from the Lists sheet -- so editing directly in
    Excel (e.g. from a phone) offers a real dropdown instead of free
    text, the exact kind of typo transaction_validator.py otherwise has
    to catch after the fact.

    Idempotent by design (clears this sheet's own data validations first,
    then re-adds) -- real Excel silently upgrades openpyxl-written
    DataValidation to its own newer extLst format on save, which openpyxl
    can't read back (a "Data Validation extension is not supported and
    will be removed" warning, then gone for good on the app's next save).
    Re-asserting these on every write (see DynamicSchema._refresh_derived_sheets())
    heals that automatically instead of the dropdowns silently vanishing
    the first time the file round-trips through real Excel."""
    ws.data_validations.dataValidation = []
    for role, (list_col, list_max) in _LIST_SOURCE.items():
        if role not in col:
            continue
        dv = DataValidation(type="list", formula1=f"={lists_sheet}!${list_col}$2:${list_col}${list_max}", allow_blank=True)
        ws.add_data_validation(dv)
        letter = get_column_letter(col[role])
        dv.add(f"{letter}2:{letter}{max_row}")


def add_category_pie(ws, category_count: int, total_col: int) -> None:
    """A pie chart of each category's yearly total (the sheet's own
    Total column) -- placed to the right of the data grid."""
    if category_count == 0:
        return
    data = Reference(ws, min_col=total_col, min_row=1, max_row=category_count + 1)
    labels = Reference(ws, min_col=1, min_row=2, max_row=category_count + 1)
    chart = PieChart()
    chart.title = "Expenses by category"
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.height, chart.width = 10, 16
    ws.add_chart(chart, f"{get_column_letter(total_col + 2)}2")


def add_monthly_trend_chart(ws, income_col: int, expense_col: int, balance_col: int) -> None:
    """A line chart of Income/Expense/Balance across the 12 month rows."""
    chart = LineChart()
    chart.title = "Income / Expense / Balance by month"
    chart.height, chart.width = 10, 20
    labels = Reference(ws, min_col=1, min_row=2, max_row=13)
    for col in (income_col, expense_col, balance_col):
        chart.add_data(Reference(ws, min_col=col, min_row=1, max_row=13), titles_from_data=True)
    chart.set_categories(labels)
    ws.add_chart(chart, f"{get_column_letter(max(income_col, expense_col, balance_col) + 2)}2")


def add_currency_trend_chart(ws, net_cols: dict[str, int]) -> None:
    """A line chart with one series per non-base currency's net monthly
    movement ({cur} Net columns) across the 12 month rows -- "all
    transactions, split by currency" in one place, next to the base-currency
    trend chart above. No-op with nothing to chart (no currency tracking)."""
    if not net_cols:
        return
    chart = LineChart()
    chart.title = "Net movement by currency (native units)"
    chart.height, chart.width = 10, 20
    labels = Reference(ws, min_col=1, min_row=2, max_row=13)
    for col in net_cols.values():
        chart.add_data(Reference(ws, min_col=col, min_row=1, max_row=13), titles_from_data=True)
    chart.set_categories(labels)
    ws.add_chart(chart, f"{get_column_letter(max(net_cols.values()) + 2)}18")
