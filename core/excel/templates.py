"""
core/excel/templates.py — generate a brand-new, blank workbook matching one
of the two real layouts this app understands, so someone without an
existing Finances_*.xlsx can start from scratch. Mirrors the *structure*
(sheets, headers, formulas) of the real files exactly, but never their
data — categories are seeded with a small generic starter set, not any
particular user's personal list.

Deliberately excluded (out of scope, see plan): the "Annual Summary"
sheet the real files also have — it's a hand-built per-category rollup
tied to specific category names, and the app itself never reads it.
"""
from __future__ import annotations

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from . import derived_sheets, schema_dynamic, styling, workbook_io
from .base import MONTH_NAMES
from .template_model import ROLE_LABELS, Template
from .schema_2025 import (
    CATEGORIES_SHEET as SCHEMA_2025_CATEGORIES_SHEET,
    COL_AMOUNT as S25_COL_AMOUNT,
    COL_CATEGORY as S25_COL_CATEGORY,
    COL_NOTES as S25_COL_NOTES,
    COL_PAYMENT as S25_COL_PAYMENT,
    COL_TYPE as S25_COL_TYPE,
    PAYMENT_HEADER_ROW as S25_PAYMENT_HEADER_ROW,
)
from .schema_2025 import DATA_START_ROW as S25_DATA_START_ROW, MAX_DATA_ROW as S25_MAX_DATA_ROW
from .schema_2026 import (
    ALLDATA_SHEET,
    ALL_COL_AMOUNT,
    ALL_COL_CATEGORY,
    ALL_COL_DATE,
    ALL_COL_NET,
    ALL_COL_TYPE,
    COL_AMOUNT as S26_COL_AMOUNT,
    COL_CATEGORY as S26_COL_CATEGORY,
    COL_DATE as S26_COL_DATE,
    COL_NET as S26_COL_NET,
    COL_NOTES as S26_COL_NOTES,
    COL_PAYMENT as S26_COL_PAYMENT,
    COL_TYPE as S26_COL_TYPE,
    LISTS_CATEGORIES_COL,
    LISTS_PAYMENT_COL,
    LISTS_SHEET,
    LISTS_TYPES_COL,
)
from .schema_2026 import DATA_START_ROW as S26_DATA_START_ROW, MAX_DATA_ROW as S26_MAX_DATA_ROW

DEFAULT_CATEGORIES = [
    "Food", "Transport", "Utilities", "Entertainment",
    "Health", "Shopping", "Subscriptions", "Crypto", "Other",
]


def _write_column(ws: Worksheet, values: list[str], start_row: int, col: int) -> None:
    for i, value in enumerate(values):
        ws.cell(row=start_row + i, column=col).value = value


def create_2025_style_workbook(path, year: int) -> None:
    """A blank workbook matching the "Simple" layout: no per-transaction
    date, Category/Type/Amount/Notes/Payment columns, one sheet per month
    with a Totals summary block, plus a flat Categories list."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for month_name in MONTH_NAMES:
        ws = wb.create_sheet(month_name)
        ws.cell(row=1, column=1).value = f"{month_name} {year}"

        header_row = S25_PAYMENT_HEADER_ROW  # = 3, same row as the other headers
        ws.cell(row=header_row, column=S25_COL_CATEGORY).value = "Category"
        ws.cell(row=header_row, column=S25_COL_TYPE).value = "Type"
        ws.cell(row=header_row, column=S25_COL_AMOUNT).value = "Amount"
        ws.cell(row=header_row, column=S25_COL_NOTES).value = "Notes"
        ws.cell(row=header_row, column=S25_COL_PAYMENT).value = "Payment type"

        d0, d1 = S25_DATA_START_ROW, S25_MAX_DATA_ROW
        b, c, e = "B", "C", "E"
        ws.cell(row=header_row, column=6).value = "Totals"
        ws.cell(row=4, column=6).value = "Total Income"
        ws.cell(row=4, column=7).value = f'=SUMIF(${b}${d0}:${b}${d1},"Income",${c}${d0}:${c}${d1})'
        ws.cell(row=5, column=6).value = "Total Expenses"
        ws.cell(row=5, column=7).value = f'=SUMIF(${b}${d0}:${b}${d1},"Expenses",${c}${d0}:${c}${d1})'
        ws.cell(row=5, column=8).value = "=G5/G4"
        ws.cell(row=6, column=6).value = "Total investments"
        ws.cell(row=6, column=7).value = f'=SUMIF($A${d0}:$A${d1},"Crypto",${c}${d0}:${c}${d1})'
        ws.cell(row=6, column=8).value = "=G6/G4"
        ws.cell(row=7, column=6).value = "Total Savings"
        ws.cell(row=7, column=7).value = (
            f'=SUMIF(${b}${d0}:${b}${d1},"Savings",${c}${d0}:${c}${d1})'
            f'-SUMIF(${b}${d0}:${b}${d1},"Cash Expense",${c}${d0}:${c}${d1})'
            f'-SUMIFS(${c}${d0}:${c}${d1},${b}${d0}:${b}${d1},"Expenses",${e}${d0}:${e}${d1},"Cash")'
        )
        ws.cell(row=7, column=8).value = "=G7/G4"
        ws.cell(row=8, column=6).value = "Net Balance"
        ws.cell(row=8, column=7).value = "=G4-G5"

    ws_cat = wb.create_sheet(SCHEMA_2025_CATEGORIES_SHEET)
    _write_column(ws_cat, DEFAULT_CATEGORIES, start_row=1, col=1)

    workbook_io.save(wb, path)


def create_2026_style_workbook(path, year: int) -> None:
    """A blank workbook matching the "Detailed" layout: per-transaction
    Date column, a running Net Change formula per row, an AllData log
    sheet, and a Lists sheet for Categories/Types/Payment options."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for month_name in MONTH_NAMES:
        ws = wb.create_sheet(month_name)
        ws.cell(row=1, column=S26_COL_DATE).value = "Date"
        ws.cell(row=1, column=S26_COL_TYPE).value = "Type"
        ws.cell(row=1, column=S26_COL_CATEGORY).value = "Category"
        ws.cell(row=1, column=S26_COL_AMOUNT).value = "Amount"
        ws.cell(row=1, column=S26_COL_NET).value = "Net Change"
        ws.cell(row=1, column=S26_COL_PAYMENT).value = "Payment type"
        ws.cell(row=1, column=S26_COL_NOTES).value = "Notes"

        d0, d1 = S26_DATA_START_ROW, S26_MAX_DATA_ROW
        ws.cell(row=2, column=9).value = "Totals"
        ws.cell(row=3, column=9).value = "Income"
        ws.cell(row=3, column=10).value = f'=SUMIF(B{d0}:B{d1},"Income",D{d0}:D{d1})'
        ws.cell(row=4, column=9).value = "Expenses"
        ws.cell(row=4, column=10).value = f'=SUMIF(B{d0}:B{d1},"Expense",D{d0}:D{d1})'
        ws.cell(row=5, column=9).value = "Invest"
        ws.cell(row=5, column=10).value = (
            f'=SUMIF($C${d0}:$C$101,"Crypto",$D${d0}:$D$101)'
            f' + SUMIF($C${d0}:$C$101,"Stocks",$D${d0}:$D$101)'
        )
        ws.cell(row=6, column=9).value = "Net Balance"
        ws.cell(row=6, column=10).value = "=J3-J4"

        ws.cell(row=8, column=17).value = "Cash"
        ws.cell(row=9, column=17).value = "Income"
        ws.cell(row=9, column=18).value = "Expense"
        ws.cell(row=9, column=19).value = "Net"
        ws.cell(row=10, column=17).value = f'=SUMIFS(D{d0}:D{d1}, B{d0}:B{d1},"To Cash",F{d0}:F{d1}, "Cash")'
        ws.cell(row=10, column=18).value = f'=SUMIFS(D{d0}:D{d1}, B{d0}:B{d1},"Expense",F{d0}:F{d1}, "Cash")'
        ws.cell(row=10, column=19).value = "=Q10-R10"

    ws_all = wb.create_sheet(ALLDATA_SHEET)
    ws_all.cell(row=1, column=ALL_COL_DATE).value = "Date"
    ws_all.cell(row=1, column=ALL_COL_TYPE).value = "Type"
    ws_all.cell(row=1, column=ALL_COL_CATEGORY).value = "Category"
    ws_all.cell(row=1, column=ALL_COL_AMOUNT).value = "Amount"
    ws_all.cell(row=1, column=ALL_COL_NET).value = "Net Change"

    ws_lists = wb.create_sheet(LISTS_SHEET)
    ws_lists.cell(row=1, column=LISTS_CATEGORIES_COL).value = "Categories"
    ws_lists.cell(row=1, column=LISTS_TYPES_COL).value = "Types"
    ws_lists.cell(row=1, column=LISTS_PAYMENT_COL).value = "Payment type"
    _write_column(ws_lists, DEFAULT_CATEGORIES, start_row=2, col=LISTS_CATEGORIES_COL)
    _write_column(ws_lists, ["Income", "Expense", "To Cash"], start_row=2, col=LISTS_TYPES_COL)
    _write_column(ws_lists, ["Cash", "Card"], start_row=2, col=LISTS_PAYMENT_COL)

    workbook_io.save(wb, path)


def create_custom_workbook(path, year: int, template: Template) -> None:
    """A blank workbook matching a user-defined Template: header row 1
    built from template.columns (whichever roles/order it chose), no data
    rows, no native Excel formulas (the app computes everything itself —
    see schema_dynamic.py), plus a Lists sheet for Categories/Types/
    Payment options."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    col_map = {role: i + 1 for i, role in enumerate(template.columns)}

    for month_name in MONTH_NAMES:
        ws = wb.create_sheet(month_name)
        for col, role in enumerate(template.columns, start=1):
            ws.cell(row=1, column=col).value = ROLE_LABELS[role]
        styling.style_transaction_sheet(ws, col_map)
        styling.add_effective_amount_column(ws, col_map)

    ws_lists = wb.create_sheet(schema_dynamic.LISTS_SHEET)
    ws_lists.cell(row=1, column=schema_dynamic.LISTS_CATEGORIES_COL).value = "Categories"
    ws_lists.cell(row=1, column=schema_dynamic.LISTS_TYPES_COL).value = "Types"
    ws_lists.cell(row=1, column=schema_dynamic.LISTS_PAYMENT_COL).value = "Payment type"
    _write_column(ws_lists, template.categories, start_row=2, col=schema_dynamic.LISTS_CATEGORIES_COL)
    _write_column(ws_lists, template.types, start_row=2, col=schema_dynamic.LISTS_TYPES_COL)
    if template.payment_types:
        _write_column(ws_lists, template.payment_types, start_row=2, col=schema_dynamic.LISTS_PAYMENT_COL)

    if template.currencies:
        ws_lists.cell(row=1, column=schema_dynamic.LISTS_CURRENCY_COL).value = "Currency"
        _write_column(ws_lists, template.currencies, start_row=2, col=schema_dynamic.LISTS_CURRENCY_COL)

        ws_lists.cell(row=1, column=schema_dynamic.LISTS_RATE_CURRENCY_COL).value = "Currency"
        ws_lists.cell(row=1, column=schema_dynamic.LISTS_RATE_VALUE_COL).value = "Rate → base"
        _write_column(ws_lists, template.currencies, start_row=2, col=schema_dynamic.LISTS_RATE_CURRENCY_COL)
        if template.base_currency in template.currencies:
            base_row = 2 + template.currencies.index(template.base_currency)
            ws_lists.cell(row=base_row, column=schema_dynamic.LISTS_RATE_VALUE_COL).value = 1.00
        # other currencies' rate cells are left blank on purpose — the user
        # fills in real rates directly in Excel; DynamicSchema.to_base_amount()
        # treats a missing rate as "don't convert" rather than crashing.

    for month_name in MONTH_NAMES:
        styling.add_dropdowns(wb[month_name], col_map)

    derived_sheets.create_sheets(wb, template, year)
    workbook_io.save(wb, path)
