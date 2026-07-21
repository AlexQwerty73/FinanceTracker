"""
core/excel/derived_sheets.py — the four sheets DynamicSchema keeps
alongside the 12 pure-transaction month sheets: All Transactions (every
row, every month, one place to scroll), Monthly Summary (one row per
month: totals + a per-currency income/expense/cash/card breakdown),
Annual Summary (Monthly Summary's own columns, summed for the year), By
Category (category x month expense grid).

Monthly Summary / Annual Summary / By Category are real Excel formulas
(SUMIFS/VLOOKUP/arithmetic — no spilling arrays, this project already
fought that fight once and moved away from it), written **once**, at file
creation (create_sheets()), and never touched again by Python — Excel
recalculates them forever after, regardless of whether a transaction was
entered through the app or typed directly into a month sheet from a phone.
All Transactions is the one exception: combining 12 variable-length sheets
into one flat list has no non-spilling formula equivalent, so it stays a
Python-maintained mirror, rebuilt on every write (refresh()) and once per
app launch (see app/window.py) — "as fresh as the last time the app
touched the file," the same limitation Schema2026's own AllData mirror
already has.
"""
from __future__ import annotations

from openpyxl.utils import get_column_letter

from . import styling
from .base import MONTH_NAMES
from .template_model import (
    ROLE_AMOUNT, ROLE_BASE_AMOUNT, ROLE_CATEGORY, ROLE_CURRENCY, ROLE_DATE, ROLE_LABELS,
    ROLE_NOTES, ROLE_PAYMENT, ROLE_RATE, ROLE_TYPE, Template,
)

ALL_TRANSACTIONS_SHEET = "All Transactions"
MONTHLY_SUMMARY_SHEET = "Monthly Summary"
ANNUAL_SUMMARY_SHEET = "Annual Summary"
BY_CATEGORY_SHEET = "By Category"

_ROLE_TO_TXKEY = {
    ROLE_DATE: "date", ROLE_TYPE: "type", ROLE_CATEGORY: "category", ROLE_AMOUNT: "amount",
    ROLE_PAYMENT: "payment_type", ROLE_CURRENCY: "currency", ROLE_NOTES: "note",
    ROLE_RATE: "rate", ROLE_BASE_AMOUNT: "base_amount",
}

_CATEGORY_ROW_HEADROOM = 15  # spare By Category rows for categories added later (app or phone)


def _summary_columns(template: Template) -> list[str]:
    cols = ["Income", "Expense", "Balance", "Invest", "Cash Δ", "Card Δ"]
    for cur in template.currencies or []:
        if cur != template.base_currency:
            cols += [f"{cur} Income", f"{cur} Expense", f"{cur} Cash Δ", f"{cur} Card Δ", f"{cur} Net"]
    return cols


def _write_monthly_summary_formulas(ws, template: Template, col_map: dict[str, int]) -> dict[str, int]:
    """Writes one SUMIFS-based row per month. Returns {label: column} so
    create_sheets() can point Annual Summary / the currency chart at the
    right columns without re-deriving the layout."""
    summary_col = {label: i for i, label in enumerate(_summary_columns(template), start=2)}
    end = styling.MAX_STYLED_ROW
    tcol, ccol, acol = (get_column_letter(col_map[r]) for r in (ROLE_TYPE, ROLE_CATEGORY, ROLE_AMOUNT))
    pcol = get_column_letter(col_map[ROLE_PAYMENT]) if ROLE_PAYMENT in col_map else None
    curcol = get_column_letter(col_map[ROLE_CURRENCY]) if ROLE_CURRENCY in col_map else None
    helper = styling.effective_amount_col(col_map)
    vcol = get_column_letter(helper) if helper else acol

    for m, month_name in enumerate(MONTH_NAMES, start=1):
        r = m + 1

        def rng(letter: str) -> str:
            return f"{month_name}!${letter}$2:${letter}${end}"

        def sumifs(value_letter: str, *criteria: tuple[str, str]) -> str:
            parts = ",".join(f'{rng(k)},"{v}"' for k, v in criteria)
            return f"SUMIFS({rng(value_letter)},{parts})"

        def put(label: str, formula: str) -> str:
            letter = get_column_letter(summary_col[label])
            ws.cell(row=r, column=summary_col[label]).value = formula
            return letter

        income_letter = put("Income", f"={sumifs(vcol, (tcol, template.income_type))}")
        expense_letter = put("Expense", f"={sumifs(vcol, (tcol, template.expense_type))}")
        balance_letter = put("Balance", f"={income_letter}{r}-{expense_letter}{r}")
        invest_terms = [sumifs(vcol, (ccol, cat)) for cat in (template.invest_categories or [])]
        put("Invest", "=" + "+".join(invest_terms) if invest_terms else "=0")
        if template.cash_in_type and pcol:
            cash_formula = f"={sumifs(vcol, (tcol, template.cash_in_type))}-{sumifs(vcol, (tcol, template.expense_type), (pcol, 'Cash'))}"
        else:
            cash_formula = "=0"
        cash_letter = put("Cash Δ", cash_formula)
        put("Card Δ", f"={balance_letter}{r}-{cash_letter}{r}")

        for cur in (template.currencies or []):
            if cur == template.base_currency:
                continue
            cur_income_letter = put(f"{cur} Income", f"={sumifs(acol, (tcol, template.income_type), (curcol, cur))}")
            cur_expense_letter = put(f"{cur} Expense", f"={sumifs(acol, (tcol, template.expense_type), (curcol, cur))}")
            if template.cash_in_type and pcol:
                cur_cash_formula = (
                    f"={sumifs(acol, (tcol, template.cash_in_type), (curcol, cur))}"
                    f"-{sumifs(acol, (tcol, template.expense_type), (pcol, 'Cash'), (curcol, cur))}"
                )
            else:
                cur_cash_formula = "=0"
            cur_cash_letter = put(f"{cur} Cash Δ", cur_cash_formula)
            cur_net_letter = put(f"{cur} Net", f"={cur_income_letter}{r}-{cur_expense_letter}{r}")
            put(f"{cur} Card Δ", f"={cur_net_letter}{r}-{cur_cash_letter}{r}")

    return summary_col


def _write_by_category_formulas(ws, template: Template, col_map: dict[str, int], max_row: int) -> None:
    end = styling.MAX_STYLED_ROW
    tcol, ccol = get_column_letter(col_map[ROLE_TYPE]), get_column_letter(col_map[ROLE_CATEGORY])
    helper = styling.effective_amount_col(col_map)
    vcol = get_column_letter(helper) if helper else get_column_letter(col_map[ROLE_AMOUNT])
    for r in range(2, max_row + 1):
        ws.cell(row=r, column=1).value = f"=Lists!A{r}"
        for c, month_name in enumerate(MONTH_NAMES, start=2):
            value_rng = f"{month_name}!${vcol}$2:${vcol}${end}"
            type_rng = f"{month_name}!${tcol}$2:${tcol}${end}"
            cat_rng = f"{month_name}!${ccol}$2:${ccol}${end}"
            ws.cell(row=r, column=c).value = f'=SUMIFS({value_rng},{type_rng},"{template.expense_type}",{cat_rng},$A{r})'
        ws.cell(row=r, column=14).value = f"=SUM(B{r}:M{r})"
        ws.cell(row=r, column=15).value = f"=IFERROR(N{r}/SUM($N$2:$N${max_row})*100,0)"


def create_sheets(wb, template: Template, year: int) -> None:
    col_map = {role: i + 1 for i, role in enumerate(template.columns)}
    ws_all = wb.create_sheet(ALL_TRANSACTIONS_SHEET)
    for col, role in enumerate(template.columns, start=1):
        ws_all.cell(row=1, column=col).value = ROLE_LABELS[role]
    styling.style_transaction_sheet(ws_all, col_map, max_row=500)

    summary_cols = _summary_columns(template)
    summary_ncols = 1 + len(summary_cols)
    ws_summary = wb.create_sheet(MONTHLY_SUMMARY_SHEET)
    ws_summary.cell(row=1, column=1).value = "Month"
    for col, label in enumerate(summary_cols, start=2):
        ws_summary.cell(row=1, column=col).value = label
    for m, month_name in enumerate(MONTH_NAMES, start=1):
        ws_summary.cell(row=m + 1, column=1).value = month_name
    styling.style_grid_sheet(ws_summary, summary_ncols, max_row=13)
    summary_col = _write_monthly_summary_formulas(ws_summary, template, col_map)
    styling.add_monthly_trend_chart(ws_summary, income_col=2, expense_col=3, balance_col=4)
    net_cols = {cur: summary_col[f"{cur} Net"] for cur in (template.currencies or []) if cur != template.base_currency}
    styling.add_currency_trend_chart(ws_summary, net_cols)

    ws_annual = wb.create_sheet(ANNUAL_SUMMARY_SHEET)
    ws_annual.cell(row=1, column=1).value = "Year"
    ws_annual.cell(row=2, column=1).value = year
    for col, label in enumerate(summary_cols, start=2):
        ws_annual.cell(row=1, column=col).value = label
    styling.style_grid_sheet(ws_annual, summary_ncols, max_row=2)
    for label, col in summary_col.items():
        letter = get_column_letter(col)
        ws_annual.cell(row=2, column=col).value = f"=SUM('{MONTHLY_SUMMARY_SHEET}'!{letter}2:{letter}13)"

    ws_cat = wb.create_sheet(BY_CATEGORY_SHEET)
    ws_cat.cell(row=1, column=1).value = "Category"
    for col, name in enumerate(MONTH_NAMES, start=2):
        ws_cat.cell(row=1, column=col).value = name[:3]
    ws_cat.cell(row=1, column=14).value = "Total"
    ws_cat.cell(row=1, column=15).value = "%"
    cat_max_row = 1 + len(template.categories) + _CATEGORY_ROW_HEADROOM
    styling.style_grid_sheet(ws_cat, 15, max_row=cat_max_row)
    for row in range(2, cat_max_row + 1):
        ws_cat.cell(row=row, column=15).number_format = '0.0"%"'
    _write_by_category_formulas(ws_cat, template, col_map, cat_max_row)
    styling.add_category_pie(ws_cat, cat_max_row - 1, total_col=14)


def _clear_rows(ws) -> None:
    for row in ws.iter_rows(min_row=2, max_row=max(ws.max_row, 2)):
        for cell in row:
            cell.value = None


def _owns_all_transactions_sheet(ws_all, schema) -> bool:
    """Whether this sheet was actually built by create_sheets() -- i.e. its
    header row matches schema's own role->column mapping exactly -- rather
    than merely happening to be *named* "All Transactions" by something
    else entirely (e.g. an older hand-built workbook). Checked every time,
    not just once, because trusting a name match alone once caused
    refresh() to clear+overwrite a same-named-but-differently-shaped sheet
    from an older file, destroying columns it didn't know about."""
    return all(ws_all.cell(row=1, column=col).value == ROLE_LABELS[role] for role, col in schema._col.items())


def refresh(wb, schema) -> bool:
    """Rebuild All Transactions from what schema.transactions_for_month()
    actually returns for each month -- the one derived sheet that can't be
    a live Excel formula (see module docstring) -- and re-assert every
    month sheet's dropdowns, which real Excel can silently break (it
    upgrades openpyxl-written DataValidation to its own newer extLst
    format on save; openpyxl can't read that back and drops it on its own
    next save -- see styling.add_dropdowns()'s docstring). Returns whether
    it actually touched anything, so the caller can skip an unnecessary
    save (and the mtime/file-watcher churn that comes with one) when there
    was nothing to do. No-ops for a workbook that predates these sheets
    (an older template/file), or whose "All Transactions" sheet wasn't
    actually built by create_sheets() (same name, different origin/shape
    -- see _owns_all_transactions_sheet()) -- both signal "not a file this
    module should be rewriting sheets in." Monthly Summary/Annual
    Summary/By Category need no Python maintenance at all -- they're real
    formulas, written once at create_sheets() time."""
    if ALL_TRANSACTIONS_SHEET not in wb.sheetnames:
        return False
    ws_all = wb[ALL_TRANSACTIONS_SHEET]
    if not _owns_all_transactions_sheet(ws_all, schema):
        return False

    for month_name in MONTH_NAMES:
        styling.add_dropdowns(wb[month_name], schema._col)

    _clear_rows(ws_all)
    row = 2
    for m in range(1, 13):
        for tx in reversed(schema.transactions_for_month(m)):  # oldest-first, chronological overall
            for role, col in schema._col.items():
                ws_all.cell(row=row, column=col).value = tx.get(_ROLE_TO_TXKEY[role])
            row += 1
    return True
