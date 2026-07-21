"""
core/excel/schema_dynamic.py — DynamicSchema: reads/writes a workbook
generated from a user-configurable Template (core/excel/template_model.py)
instead of a hardcoded column layout. One instance per (file, template)
pair, constructed by the registry once it resolves a year's registered
template id to something other than the two fixed presets.

Column positions come from template.columns (an ordered list of roles)
rather than module-level constants — everything else (row scanning,
formula-free Amount parsing, save/load) reuses the exact same helpers
Schema2025/Schema2026 already use, so this is deliberately close to a
parameterized version of schema_2025.py's simpler (no-AllData-mirror)
approach.

Scope, per the approved plan: no Excel-native Totals/SUMIF block, no
AllData-style mirror log, cash-out is payment-based only (no legacy
type-based hack). Investment tracking mirrors Schema2025/Schema2026's
category-membership approach (a fixed set of category names counts
toward "invest" regardless of Income/Expense type), just sourced from
template.invest_categories instead of a module constant.
"""
from __future__ import annotations

from datetime import date as Date, datetime

from openpyxl.utils import get_column_letter

from .. import rate_history
from . import derived_sheets, transaction_reader, workbook_io
from ._formula import amount_value
from ._rows import find_empty_row
from .base import (
    CategoryExistsError, CategoryInUseError, MONTH_NAMES, SheetFullError, TransactionNotFoundError, YearSchema,
)
from .template_model import (
    ROLE_AMOUNT, ROLE_BASE_AMOUNT, ROLE_CATEGORY, ROLE_CURRENCY, ROLE_DATE, ROLE_NOTES, ROLE_PAYMENT,
    ROLE_RATE, ROLE_TYPE, Template,
)
from .transaction_reader import DATA_START_ROW, MAX_DATA_ROW

LISTS_SHEET = "Lists"
LISTS_CATEGORIES_COL, LISTS_TYPES_COL, LISTS_PAYMENT_COL = 1, 2, 3
LISTS_CURRENCY_COL = 4  # currency codes list (dropdown source)
LISTS_RATE_CURRENCY_COL, LISTS_RATE_VALUE_COL = 8, 9  # small Currency -> Rate table, columns H:I


def _base_amount_formula(row: int, amount_col: int, rate_col: int) -> str:
    """=IF(<rate cell>="","",<amount cell>*<rate cell>) -- a real, visible
    Excel formula rather than a literal number, so Amount(base currency)
    always shows its own arithmetic and self-recomputes the instant the
    Rate cell next to it is edited, whether that edit came from the app
    or was typed directly into Excel. Blank whenever Rate is blank,
    matching the old "None means not yet resolved" convention. Trade-off,
    accepted deliberately: since this is now a cell-reference formula
    (not a bare number), amount_value() can't evaluate it, so the app's
    own convert_transaction()/month_summary() can no longer trust this
    cell directly -- they always fall through to a fresh Python-side
    resolution instead (see their own docstrings), which is still
    correct, just independent of whatever a manually-typed Rate shows in
    Excel until the app itself resolves/caches that same rate."""
    amount_letter, rate_letter = get_column_letter(amount_col), get_column_letter(rate_col)
    return f'=IF({rate_letter}{row}="","",{amount_letter}{row}*{rate_letter}{row})'


def rate_from_amounts(native_amount: float, base_amount: float) -> float | None:
    """base_amount / native_amount -- the rate implied by editing the
    converted (base-currency) amount directly instead of typing a raw
    rate (e.g. "the bank actually took 250 Kč off my card for this 10
    USD purchase"). None if native_amount is 0 (undefined). Shared by
    TransactionDialog and TransactionsPage so this formula exists in
    exactly one place."""
    if native_amount == 0:
        return None
    return base_amount / native_amount


class DynamicSchema(YearSchema):
    def __init__(self, file_path, year: int, template: Template):
        super().__init__(file_path, year)
        self.template = template
        self.EXPENSE_TYPE = template.expense_type
        self.INCOME_TYPE = template.income_type
        self.CASH_IN_TYPE = template.cash_in_type
        self.HAS_DAILY_DATES = template.has_daily_dates()
        self.INVEST_CATEGORIES = set(template.invest_categories or [])
        self._col = {role: i + 1 for i, role in enumerate(template.columns)}

    def get_payment_types(self) -> list[str] | None:
        if ROLE_PAYMENT not in self._col:
            return None
        return list(self.template.payment_types or [])

    def get_currencies(self) -> list[str] | None:
        if ROLE_CURRENCY not in self._col:
            return None
        return list(self.template.currencies or [])

    def get_base_currency(self) -> str | None:
        if ROLE_CURRENCY not in self._col:
            return None
        return self.template.base_currency

    def to_base_amount(self, amount: float, currency: str | None, date: Date | None = None) -> float:
        if ROLE_CURRENCY not in self._col:
            return amount
        rate = self.resolve_rate(currency, date)
        return amount * rate if rate is not None else amount

    def resolve_rate(self, currency: str | None, date: Date | None) -> float | None:
        if not currency:
            return None
        if currency == self.template.base_currency:
            return 1.0
        if date is not None:
            historical_date = date.date() if hasattr(date, "date") else date
            historical = rate_history.get_rate(currency, historical_date)
            if historical is not None:
                return historical
        return self._read_rates().get(currency)

    def convert_transaction(self, tx: dict) -> float:
        """The base-currency amount for one transaction dict (as returned
        by transactions_for_month()) -- trusts the row's own persisted
        base_amount cell when there is one (the table is the database: an
        error actually written there must show up here too, not get
        silently recomputed away), else falls back to to_base_amount()
        (historical/manual rate, pinned to the row's own date) -- Income
        and Expense are treated identically: a rate pinned to when the
        money actually changed hands, exactly like a bank statement. (An
        earlier version of this app let foreign-currency Income "float" at
        the latest known rate instead, on the theory that unexchanged
        income isn't realized yet -- reverted after the user found the
        real flaw: a past month's Income total would silently reshuffle
        itself every time the exchange rate moved, even after that money
        had already been spent at the old rate. "What's this worth if I
        exchanged it today" is a separate, legitimate question -- that's
        what CurrenciesPage's "Total right now" already answers
        independently, using today's rate against today's actual balance,
        not by reinterpreting historical transactions.)"""
        if tx.get("base_amount") is not None:
            return tx["base_amount"]
        amount, currency = tx.get("amount") or 0.0, tx.get("currency")
        return self.to_base_amount(amount, currency, tx.get("date"))

    def get_rates(self) -> dict[str, float] | None:
        if ROLE_CURRENCY not in self._col:
            return None
        return self._read_rates()

    def _read_rates(self) -> dict[str, float]:
        wb = workbook_io.load(self.file_path, data_only=False)
        ws = wb[LISTS_SHEET]
        rates: dict[str, float] = {}
        row = 2
        while True:
            code = ws.cell(row=row, column=LISTS_RATE_CURRENCY_COL).value
            if code is None:
                break
            val = ws.cell(row=row, column=LISTS_RATE_VALUE_COL).value
            if isinstance(val, (int, float)):
                rates[str(code).strip()] = float(val)
            row += 1
        return rates

    def update_current_rate(self, currency: str, rate: float) -> None:
        """Overwrite Lists!H:I's "current rate" cell for `currency` --
        called after a background sync fetches a fresh rate, so the
        file's own at-a-glance rate table stays honest too, not just the
        per-transaction Rate/Amount(base) cells (see
        feedback_excel_immediate_sync.md). No-ops if `currency` isn't
        already listed there."""
        wb = workbook_io.load(self.file_path, data_only=False)
        ws = wb[LISTS_SHEET]
        row = 2
        while ws.cell(row=row, column=LISTS_RATE_CURRENCY_COL).value is not None:
            if str(ws.cell(row=row, column=LISTS_RATE_CURRENCY_COL).value).strip() == currency:
                workbook_io.invalidate(self.file_path)
                ws.cell(row=row, column=LISTS_RATE_VALUE_COL).value = rate
                workbook_io.save(wb, self.file_path)
                return
            row += 1

    # ── Lists sheet (Categories/Types/Payment) ──────────────────────────

    def get_categories(self) -> list[str]:
        wb = workbook_io.load(self.file_path, data_only=False)
        return self._read_list_column(wb[LISTS_SHEET], LISTS_CATEGORIES_COL)

    def get_types(self) -> list[str]:
        wb = workbook_io.load(self.file_path, data_only=False)
        return self._read_list_column(wb[LISTS_SHEET], LISTS_TYPES_COL)

    @staticmethod
    def _read_list_column(ws, col: int) -> list[str]:
        values: list[str] = []
        row = 2  # row 1 is the "Categories"/"Types"/"Payment type" header
        while True:
            v = ws.cell(row=row, column=col).value
            if v is None:
                break
            values.append(str(v).strip())
            row += 1
        return values

    def _refresh_derived_sheets(self) -> None:
        """Rebuild All Transactions (derived_sheets.py) from what's now on
        the month sheets -- Monthly/Annual Summary and By Category need no
        rebuilding here, they're live Excel formulas written once at file
        creation. A separate load+save from whatever mutation just
        happened, so it always rebuilds from what's actually on disk. Saves
        only if refresh() actually changed something -- e.g. a workbook
        that predates these sheets, or one whose same-named sheet wasn't
        actually built by create_sheets(), must be left byte-for-byte
        untouched, not silently re-saved with no real change."""
        wb = workbook_io.load(self.file_path, data_only=False)
        workbook_io.invalidate(self.file_path)
        if derived_sheets.refresh(wb, self):
            workbook_io.save(wb, self.file_path)

    def add_category(self, name: str) -> None:
        wb = workbook_io.load(self.file_path, data_only=False)
        workbook_io.invalidate(self.file_path)
        ws = wb[LISTS_SHEET]
        row = 2
        while ws.cell(row=row, column=LISTS_CATEGORIES_COL).value is not None:
            if str(ws.cell(row=row, column=LISTS_CATEGORIES_COL).value).strip().lower() == name.lower():
                raise CategoryExistsError(f'Category "{name}" already exists.')
            row += 1
        ws.cell(row=row, column=LISTS_CATEGORIES_COL).value = name
        workbook_io.save(wb, self.file_path)
        self._refresh_derived_sheets()

    def rename_category(self, old_name: str, new_name: str) -> int:
        wb = workbook_io.load(self.file_path, data_only=False)
        workbook_io.invalidate(self.file_path)
        ws_lists = wb[LISTS_SHEET]

        row = 2
        rename_row = None
        while ws_lists.cell(row=row, column=LISTS_CATEGORIES_COL).value is not None:
            v = str(ws_lists.cell(row=row, column=LISTS_CATEGORIES_COL).value).strip()
            if new_name != old_name and v.lower() == new_name.lower():
                raise CategoryExistsError(f'Category "{new_name}" already exists.')
            if v == old_name:
                rename_row = row
            row += 1
        if rename_row is not None:
            ws_lists.cell(row=rename_row, column=LISTS_CATEGORIES_COL).value = new_name

        count = 0
        cat_col = self._col[ROLE_CATEGORY]
        for month_name in MONTH_NAMES:
            ws = wb[month_name]
            for r in range(DATA_START_ROW, MAX_DATA_ROW + 1):
                if ws.cell(row=r, column=cat_col).value == old_name:
                    ws.cell(row=r, column=cat_col).value = new_name
                    count += 1

        workbook_io.save(wb, self.file_path)
        self._refresh_derived_sheets()
        return count

    def delete_category(self, name: str) -> None:
        wb = workbook_io.load(self.file_path, data_only=False)
        workbook_io.invalidate(self.file_path)
        ws_lists = wb[LISTS_SHEET]
        if name not in self._read_list_column(ws_lists, LISTS_CATEGORIES_COL):
            raise ValueError(f'Category "{name}" does not exist.')

        cat_col = self._col[ROLE_CATEGORY]
        count = 0
        for month_name in MONTH_NAMES:
            ws = wb[month_name]
            for r in range(DATA_START_ROW, MAX_DATA_ROW + 1):
                if ws.cell(row=r, column=cat_col).value == name:
                    count += 1
        if count > 0:
            raise CategoryInUseError(
                f'Category "{name}" is used by {count} transaction(s) — merge it into another category first.'
            )

        self._remove_list_row(ws_lists, LISTS_CATEGORIES_COL, name)
        workbook_io.save(wb, self.file_path)

    def merge_category(self, source: str, target: str) -> int:
        if source == target:
            raise ValueError("Source and target categories must be different.")
        wb = workbook_io.load(self.file_path, data_only=False)
        workbook_io.invalidate(self.file_path)
        ws_lists = wb[LISTS_SHEET]
        existing = self._read_list_column(ws_lists, LISTS_CATEGORIES_COL)
        if source not in existing:
            raise ValueError(f'Category "{source}" does not exist.')
        if target not in existing:
            raise ValueError(f'Category "{target}" does not exist.')

        cat_col = self._col[ROLE_CATEGORY]
        count = 0
        for month_name in MONTH_NAMES:
            ws = wb[month_name]
            for r in range(DATA_START_ROW, MAX_DATA_ROW + 1):
                if ws.cell(row=r, column=cat_col).value == source:
                    ws.cell(row=r, column=cat_col).value = target
                    count += 1

        self._remove_list_row(ws_lists, LISTS_CATEGORIES_COL, source)
        workbook_io.save(wb, self.file_path)
        self._refresh_derived_sheets()
        return count

    def reorder_categories(self, new_order: list[str]) -> None:
        if sorted(new_order) != sorted(self.get_categories()):
            raise ValueError("New order must contain exactly the same categories.")
        wb = workbook_io.load(self.file_path, data_only=False)
        workbook_io.invalidate(self.file_path)
        ws_lists = wb[LISTS_SHEET]
        for i, name in enumerate(new_order):
            ws_lists.cell(row=2 + i, column=LISTS_CATEGORIES_COL).value = name
        workbook_io.save(wb, self.file_path)

    @staticmethod
    def _remove_list_row(ws, col: int, name: str) -> None:
        """Shift every row after `name`'s row up by one — Lists columns are
        packed, no-gap lists (_read_list_column reads until the first blank
        cell), so removing an entry means closing the gap, not just
        blanking it."""
        row = 2
        target_row = None
        while ws.cell(row=row, column=col).value is not None:
            if str(ws.cell(row=row, column=col).value).strip() == name:
                target_row = row
                break
            row += 1
        if target_row is None:
            return
        r = target_row
        while ws.cell(row=r + 1, column=col).value is not None:
            ws.cell(row=r, column=col).value = ws.cell(row=r + 1, column=col).value
            r += 1
        ws.cell(row=r, column=col).value = None

    # ── internal: operate on an already-open workbook, no load/save ────────

    def _write_transaction(
        self, wb, date: Date, type_: str, category: str, amount: float,
        payment_type: str | None, note: str, currency: str | None = None, rate_override: float | None = None,
    ) -> None:
        ws = wb[MONTH_NAMES[date.month - 1]]
        row = find_empty_row(ws, DATA_START_ROW, MAX_DATA_ROW, self._col[ROLE_CATEGORY])
        if row is None:
            raise SheetFullError(
                f"{MONTH_NAMES[date.month - 1]} {date.year} is full "
                f"(max {MAX_DATA_ROW - DATA_START_ROW + 1} transactions)."
            )

        if ROLE_DATE in self._col:
            ws.cell(row=row, column=self._col[ROLE_DATE]).value = datetime(date.year, date.month, date.day)
        ws.cell(row=row, column=self._col[ROLE_CATEGORY]).value = category
        ws.cell(row=row, column=self._col[ROLE_TYPE]).value = type_
        ws.cell(row=row, column=self._col[ROLE_AMOUNT]).value = amount
        if ROLE_PAYMENT in self._col:
            ws.cell(row=row, column=self._col[ROLE_PAYMENT]).value = payment_type
        if ROLE_CURRENCY in self._col:
            ws.cell(row=row, column=self._col[ROLE_CURRENCY]).value = currency
        if ROLE_NOTES in self._col:
            ws.cell(row=row, column=self._col[ROLE_NOTES]).value = note
        if rate_override is not None:
            rate = rate_override
            # A manual override (e.g. the real rate a bank card charged,
            # markup included) is the best-known rate for this exact day
            # from here on — persist it to rate_history.json so a later
            # background sync doesn't silently overwrite it with the
            # "official" rate, and so resolve_rate()/to_base_amount() pick
            # it up everywhere else too. Deliberately NOT gated on whether
            # this template even has Rate/Amount(base) columns -- the cache
            # is what every conversion actually reads from; the columns are
            # just this template's *visible* copy of the same fact.
            if currency and currency != self.template.base_currency:
                historical_date = date.date() if hasattr(date, "date") else date
                rate_history.set_rate(currency, historical_date, rate)
        else:
            rate = self.resolve_rate(currency, date)
        if ROLE_RATE in self._col:
            ws.cell(row=row, column=self._col[ROLE_RATE]).value = rate
        if ROLE_BASE_AMOUNT in self._col:
            if ROLE_RATE in self._col:
                ws.cell(row=row, column=self._col[ROLE_BASE_AMOUNT]).value = _base_amount_formula(
                    row, self._col[ROLE_AMOUNT], self._col[ROLE_RATE]
                )
            else:
                ws.cell(row=row, column=self._col[ROLE_BASE_AMOUNT]).value = amount * rate if rate is not None else None

    def _clear_transaction(self, wb, tx: dict) -> None:
        ws = wb[tx["month"]]
        row = tx["_row"]
        if not (
            ws.cell(row=row, column=self._col[ROLE_TYPE]).value == tx["type"]
            and ws.cell(row=row, column=self._col[ROLE_CATEGORY]).value == tx["category"]
        ):
            raise TransactionNotFoundError(
                "This transaction no longer matches what's in the file (it may "
                "have changed externally). Refresh the dashboard and try again."
            )
        for col in self._col.values():
            ws.cell(row=row, column=col).value = None

    # ── public API ──────────────────────────────────────────────────────

    def add_transaction(
        self, date: Date, type_: str, category: str, amount: float,
        payment_type: str | None, note: str, currency: str | None = None, rate: float | None = None,
    ) -> None:
        wb = workbook_io.load(self.file_path, data_only=False)
        workbook_io.invalidate(self.file_path)
        self._write_transaction(wb, date, type_, category, amount, payment_type, note, currency, rate)
        workbook_io.save(wb, self.file_path)
        self._refresh_derived_sheets()

    def update_transaction(
        self, tx: dict, date: Date, type_: str, category: str, amount: float,
        payment_type: str | None, note: str, currency: str | None = None, rate: float | None = None,
    ) -> None:
        wb = workbook_io.load(self.file_path, data_only=False)
        workbook_io.invalidate(self.file_path)
        self._clear_transaction(wb, tx)
        self._write_transaction(wb, date, type_, category, amount, payment_type, note, currency, rate)
        workbook_io.save(wb, self.file_path)
        self._refresh_derived_sheets()

    def delete_transaction(self, tx: dict) -> None:
        wb = workbook_io.load(self.file_path, data_only=False)
        workbook_io.invalidate(self.file_path)
        self._clear_transaction(wb, tx)
        workbook_io.save(wb, self.file_path)
        self._refresh_derived_sheets()

    def refresh_converted_amounts(self, currency: str, date: Date, rate: float) -> int:
        """Rewrite the Rate/Amount (base currency) cells for every
        transaction on `date` in `currency` with `rate` — called after
        RateSyncWorker fetches a better (e.g. real historical, replacing an
        earlier current-table fallback) rate for that exact (currency,
        date) pair, so cells already written keep matching what the app
        itself now believes is correct. No-ops if this template has
        neither column. Returns the number of rows updated."""
        if ROLE_RATE not in self._col and ROLE_BASE_AMOUNT not in self._col:
            return 0
        date_col = self._col.get(ROLE_DATE)
        currency_col = self._col.get(ROLE_CURRENCY)
        if date_col is None or currency_col is None:
            return 0

        wb = workbook_io.load(self.file_path, data_only=False)
        workbook_io.invalidate(self.file_path)
        ws = wb[MONTH_NAMES[date.month - 1]]
        amount_col = self._col[ROLE_AMOUNT]
        updated = 0
        for row in range(DATA_START_ROW, MAX_DATA_ROW + 1):
            cell_date = ws.cell(row=row, column=date_col).value
            if cell_date is None:
                continue
            d = cell_date.date() if hasattr(cell_date, "date") else cell_date
            if d != date or ws.cell(row=row, column=currency_col).value != currency:
                continue
            amount = amount_value(ws.cell(row=row, column=amount_col).value) or 0
            if ROLE_RATE in self._col:
                ws.cell(row=row, column=self._col[ROLE_RATE]).value = rate
            if ROLE_BASE_AMOUNT in self._col:
                if ROLE_RATE in self._col:
                    ws.cell(row=row, column=self._col[ROLE_BASE_AMOUNT]).value = _base_amount_formula(
                        row, amount_col, self._col[ROLE_RATE]
                    )
                else:
                    ws.cell(row=row, column=self._col[ROLE_BASE_AMOUNT]).value = amount * rate
            updated += 1

        if updated:
            workbook_io.save(wb, self.file_path)
        return updated

    def month_summary(self, month: int) -> dict:
        wb = workbook_io.load(self.file_path, data_only=False)
        ws = wb[MONTH_NAMES[month - 1]]
        type_col, amount_col = self._col[ROLE_TYPE], self._col[ROLE_AMOUNT]
        cat_col = self._col[ROLE_CATEGORY]
        payment_col = self._col.get(ROLE_PAYMENT)
        currency_col = self._col.get(ROLE_CURRENCY)
        date_col = self._col.get(ROLE_DATE)
        base_amount_col = self._col.get(ROLE_BASE_AMOUNT)
        rates = self._read_rates() if currency_col else None

        income = expense = invest = cash = 0.0
        for row in range(DATA_START_ROW, MAX_DATA_ROW + 1):
            t = ws.cell(row=row, column=type_col).value
            if t is None:
                continue
            amount = amount_value(ws.cell(row=row, column=amount_col).value) or 0
            category = ws.cell(row=row, column=cat_col).value
            payment = ws.cell(row=row, column=payment_col).value if payment_col else None
            if currency_col:
                currency = ws.cell(row=row, column=currency_col).value
                # Trust a persisted Amount(base) cell directly when THIS row
                # actually has one (see refresh_converted_amounts()/
                # _write_transaction) -- the table is the database. But a
                # template merely *having* the column doesn't mean every row
                # does: a row entered directly in Excel (e.g. from a phone,
                # no Rate/Amount(base) typed in) or migrated under an older
                # rule leaves it blank, and must still fall through to a
                # fresh historical/current-table resolution below -- same
                # rule convert_transaction() already gets right. An `elif
                # base_amount_col:` here (checking only whether the column
                # exists, not this row's own cell) was a real bug: any
                # foreign-currency row with a blank Amount(base) silently
                # counted as 1:1 unconverted instead of being resolved.
                base_amount = amount_value(ws.cell(row=row, column=base_amount_col).value) if base_amount_col else None
                if base_amount is not None:
                    amount = base_amount
                elif currency and currency != self.template.base_currency:
                    date_val = ws.cell(row=row, column=date_col).value if date_col else None
                    d = date_val.date() if hasattr(date_val, "date") else date_val
                    historical = rate_history.get_rate(currency, d) if d is not None else None
                    rate = historical if historical is not None else rates.get(currency)
                    amount = amount * rate if rate is not None else amount

            if self.is_income_type(t):
                income += amount
            elif self.is_expense_type(t):
                expense += amount
            if category in self.INVEST_CATEGORIES:
                invest += amount

            if self.is_cash_in_type(t):
                cash += amount
            elif self.is_expense_type(t) and payment == "Cash":
                cash -= amount

        balance = income - expense
        has_tracking = self.has_cash_tracking()
        return {
            "income": income,
            "expense": expense,
            "invest": invest,
            "balance": balance,
            "cash": cash if has_tracking else None,
            "card": (balance - cash) if has_tracking else None,
        }

    def currency_breakdown(self, month: int) -> dict[str, dict[str, float]]:
        """{currency: {"income", "expense", "cash", "card"}} in NATIVE
        units, one entry per tracked currency other than the base -- same
        classification month_summary() uses, just kept per-currency
        instead of converted. Feeds derived_sheets.py's Monthly/Annual
        Summary sheets."""
        base = self.get_base_currency()
        result = {cur: {"income": 0.0, "expense": 0.0, "cash": 0.0, "card": 0.0}
                   for cur in self.get_currencies() or [] if cur != base}
        for tx in self.transactions_for_month(month):
            bucket = result.get(tx.get("currency"))
            if bucket is None:
                continue
            t, amount, payment = tx.get("type"), tx.get("amount") or 0, tx.get("payment_type")
            if self.is_income_type(t):
                bucket["income"] += amount
            elif self.is_expense_type(t):
                bucket["expense"] += amount
            if self.is_cash_in_type(t):
                bucket["cash"] += amount
            elif self.is_expense_type(t) and payment == "Cash":
                bucket["cash"] -= amount
        for bucket in result.values():
            bucket["card"] = (bucket["income"] - bucket["expense"]) - bucket["cash"]
        return result

    def transactions_for_month(self, month: int) -> list[dict]:
        month_name = MONTH_NAMES[month - 1]
        wb = workbook_io.load(self.file_path, data_only=False)
        return transaction_reader.read_month(wb[month_name], self._col, month_name)
